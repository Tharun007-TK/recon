import pandas as pd
import numpy as np
import io
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# Colour palette  —  "Midnight Teal" theme
# ─────────────────────────────────────────────
CLR_MIDNIGHT  = "FF0F2544"   # deep midnight navy  → section group bg
CLR_TEAL      = "FF00796B"   # rich teal           → sub-header / label bg
CLR_GOLD      = "FFFFC107"   # warm amber-gold     → label text in row-1
CLR_CYAN      = "FF80DEEA"   # soft cyan           → value text in row-1
CLR_WHITE     = "FFFFFFFF"
CLR_BLACK     = "FF000000"

# Row-status fill colours (used in data rows)
CLR_ROW_MATCHED   = "FFE0F7F0"   # mint green   → Matched
CLR_ROW_PARTIAL   = "FFFFF3E0"   # warm amber   → Partially Matched
CLR_ROW_BOOKS     = "FFE3F2FD"   # sky blue     → In Books Only
CLR_ROW_2B        = "FFEDE7F6"   # soft lavender→ In 2B Only
CLR_ROW_ALT       = "FFF5F5F5"   # light silver → alternating rows (supplier sheet)

# Back-compat aliases so every _hdr / _write_info_rows call below stays clean
CLR_SALMON  = CLR_TEAL      # was salmon,  now teal
CLR_STEEL   = CLR_MIDNIGHT  # was steel,   now midnight
CLR_LABEL   = CLR_GOLD      # was salmon,  now gold
CLR_VALUE   = CLR_CYAN      # was steel,   now cyan


def _fill(hex_rgb):
    return PatternFill("solid", fgColor=hex_rgb)


def _font(bold=False, color=CLR_BLACK, size=10):
    return Font(bold=bold, color=color, size=size)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center")


def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


# ─────────────────────────────────────────────
# Invoice-number normaliser  (THE BUG FIX)
# ─────────────────────────────────────────────
import re as _re

def _normalize_inv(s) -> str:
    """Return a canonical invoice-number key used ONLY for matching.

    Root cause fixed here:
        Tally exports use  '/'  as a separator  (e.g. "2025-26/011")
        GSTR-2B exports use '-' as a separator  (e.g. "2025-26-011")
    These are the same invoice but a plain string-equality test fails.

    Strategy:
      1. Strip surrounding whitespace.
      2. Upper-case (handles any mixed-case slugs).
      3. Collapse all '/' -> '-' so both sides share one canonical form.
      4. De-duplicate consecutive dashes introduced by step 3
         (e.g. "ABC//001" -> "ABC-001", not "ABC--001").
    The original value is ALWAYS preserved in the DataFrame for display.
    """
    if not isinstance(s, str):
        s = str(s) if s is not None else ""
    s = s.strip().upper()
    s = s.replace("/", "-")
    s = _re.sub(r"-{2,}", "-", s)
    return s


# ─────────────────────────────────────────────
# Shared header writer (rows 1-2 of every sheet)
# ─────────────────────────────────────────────
def _write_info_rows(ws, buyer_name, gstin, period, total_docs,
                     matched, partially_matched, in_books_only, in_2b_only,
                     row1_label="Total Documents →"):
    """Write the 2-row meta-info banner that appears at the top of every sheet."""
    now_str = datetime.now().strftime("%d/%m/%Y %I:%M %p")

    # Row 1: buyer / GSTIN / period / generated-on
    pairs_r1 = [
        ("Buyer Name →", buyer_name),
        ("GSTIN →", gstin),
        ("Period/F.Y. ->", period),
        ("Generated On →", now_str),
    ]
    col = 1
    for lbl, val in pairs_r1:
        c = ws.cell(row=1, column=col, value=lbl)
        c.font = _font(bold=True, color=CLR_LABEL)
        col += 1
        c = ws.cell(row=1, column=col, value=val)
        c.font = _font(bold=True, color=CLR_VALUE)
        col += 1

    # Row 2: document counts
    pairs_r2 = [
        (row1_label, total_docs),
        ("Matched →", matched),
        ("Partially Matched →", partially_matched),
        ("In Books Only →", in_books_only),
        ("In 2B Only →", in_2b_only),
    ]
    col = 1
    for lbl, val in pairs_r2:
        c = ws.cell(row=2, column=col, value=lbl)
        c.font = _font(bold=True)
        col += 1
        c = ws.cell(row=2, column=col, value=val)
        c.font = _font(bold=True)
        col += 1


# ─────────────────────────────────────────────
# Invoice-Wise sheet builder
# ─────────────────────────────────────────────
def _build_invoice_wise_sheet(ws, records, buyer_name, gstin, period,
                               matched_count, partial_count, books_only_count, gst_only_count):
    """
    Mirrors the 'Invoice Wise 2B(B2B_CNDN_ECO)' layout:
      Row 1  – buyer info banner
      Row 2  – counts banner
      Row 3  – section label "B2B (including CN/DN/ECO)"  [blue bg, full row]
      Row 4  – group headers: Supplier Name | GSTIN | Books … | GSTR 2B … | Difference …
      Row 5  – sub-column headers
      Row 6+ – data
    """
    total_docs = len(records)
    _write_info_rows(ws, buyer_name, gstin, period, total_docs,
                     matched_count, partial_count, books_only_count, gst_only_count)

    # ── Row 3: section label ──────────────────
    LAST_COL = 34   # AR  (we use 34 columns total)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=LAST_COL)
    c = ws.cell(row=3, column=1, value="B2B (including CN/DN/ECO)")
    c.fill = _fill(CLR_STEEL)
    c.font = _font(bold=True, color=CLR_WHITE)

    # ── Row 4: group headers ──────────────────
    # A4:A5 – Supplier Name
    ws.merge_cells("A4:A5")
    _hdr(ws, 4, 1, "Supplier Name", CLR_SALMON)

    # B4:B5 – GSTIN
    ws.merge_cells("B4:B5")
    _hdr(ws, 4, 2, "GSTIN", CLR_SALMON)

    # C4:M4 – Books  (11 cols: C..M)
    ws.merge_cells("C4:M4")
    _hdr(ws, 4, 3, "Books", CLR_STEEL)

    # N4:AA4 – GSTR 2B  (14 cols: N..AA)
    ws.merge_cells("N4:AA4")
    _hdr(ws, 4, 14, "GSTR 2B", CLR_STEEL)

    # AB4:AF4 – Difference  (5 cols)
    ws.merge_cells("AB4:AF4")
    _hdr(ws, 4, 28, "Difference", CLR_STEEL)

    # AG4:AH4 – (part of difference / status area, keep blank group)
    ws.merge_cells("AG4:AH4")
    _hdr(ws, 4, 33, "", CLR_STEEL)

    # ── Row 5: sub-column headers ─────────────
    # NOTE: cols 1 (A) and 2 (B) are spanned by A4:A5 and B4:B5 merges —
    # writing to those cells in row 5 would hit a MergedCell (read-only).
    # They are already labelled in row 4, so we skip them here.
    sub_headers = [
        # Books sub-cols (cols 3-13, no vertical merge from row 4)
        (3,  "Document Type"),
        (4,  "Document Number"),
        (5,  "Document Date"),
        (6,  "Taxable Value"),
        (7,  "IGST"),
        (8,  "CGST + SGST"),
        (9,  "Cess"),
        (10, "Document Value"),
        (11, "RCM"),
        (12, "ITC Availability"),
        (13, "Books Period"),
        # GSTR 2B sub-cols
        (14, "Document Type"),
        (15, "Document Number"),
        (16, "Document Date"),
        (17, "Taxable Value"),
        (18, "IGST"),
        (19, "CGST + SGST"),
        (20, "Cess"),
        (21, "Document Value"),
        (22, "RCM"),
        (23, "ITC Availability"),
        (24, "2A Period"),
        (25, "2B Period"),
        (26, "IRN Number"),
        (27, "IRN Date"),
        # Difference sub-cols
        (28, "Taxable Value Diff."),
        (29, "IGST Diff."),
        (30, "CGST + SGST Diff."),
        (31, "Cess Diff."),
        (32, "Document Value Diff."),
        (33, "Match Status"),
        (34, "Remarks"),
    ]
    for col, label in sub_headers:
        _hdr(ws, 5, col, label, CLR_SALMON)

    # ── Column widths ─────────────────────────
    col_widths = {1: 35, 2: 20, 4: 20, 5: 20, 6: 20, 11: 20,
                  15: 20, 16: 20, 17: 20, 22: 20, 26: 20,
                  27: 20, 28: 20, 33: 22, 34: 22}
    for c_idx in range(1, LAST_COL + 1):
        w = col_widths.get(c_idx, 13)
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    # ── Freeze panes ──────────────────────────
    ws.freeze_panes = "B6"

    # ── Data rows ────────────────────────────
    def _safe(v, default="-"):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return default
        if str(v).strip() in ("", "nan", "NaT", "None", "NaN"):
            return default
        return v

    def _num(v):
        try:
            return round(float(v), 2) if v not in (None, "-", "") and not (isinstance(v, float) and np.isnan(v)) else 0
        except (ValueError, TypeError):
            return 0

    for data_row in records:
        row_num = ws.max_row + 1
        status = data_row.get("Match Status", "")

        # Map our reconcile fields → the 34-column layout
        row_vals = [
            _safe(data_row.get("party_name_tally") or data_row.get("party_name_gst")),  # A Supplier Name
            _safe(data_row.get("gstin_tally") or data_row.get("gstin_gst")),             # B GSTIN
            # Books (C-M)
            _safe(data_row.get("doc_type_books", "INVOICE")),
            _safe(data_row.get("invoice_no")),
            _safe(data_row.get("invoice_date_tally")),
            _num(data_row.get("taxable_value_tally")),
            _num(data_row.get("igst_tally", 0)),
            _num(data_row.get("cgst_sgst_tally", data_row.get("tax_amount_tally", 0))),
            0,   # Cess
            _num(data_row.get("invoice_value_tally",
                              (_num(data_row.get("taxable_value_tally")) +
                               _num(data_row.get("tax_amount_tally", 0))))),
            _safe(data_row.get("rcm_books", "-")),
            _safe(data_row.get("itc_books", "-")),
            _safe(data_row.get("books_period", "-")),
            # GSTR 2B (N-AA)
            _safe(data_row.get("doc_type_gst", "INVOICE")),
            _safe(data_row.get("invoice_no_gst") or data_row.get("invoice_no")),
            _safe(data_row.get("invoice_date_gst")),
            _num(data_row.get("taxable_value_gst")),
            _num(data_row.get("igst_gst", 0)),
            _num(data_row.get("cgst_sgst_gst", data_row.get("tax_amount_gst", 0))),
            0,   # Cess
            _num(data_row.get("invoice_value_gst",
                              (_num(data_row.get("taxable_value_gst")) +
                               _num(data_row.get("tax_amount_gst", 0))))),
            _safe(data_row.get("rcm_gst", "No")),
            _safe(data_row.get("itc_gst", "Yes")),
            _safe(data_row.get("period_2a", "-")),
            _safe(data_row.get("period_2b", "-")),
            _safe(data_row.get("irn", "-")),
            _safe(data_row.get("irn_date", "-")),
            # Difference (AB-AF)
            _num(data_row.get("taxable_value_tally", 0)) - _num(data_row.get("taxable_value_gst", 0)),
            _num(data_row.get("igst_tally", 0)) - _num(data_row.get("igst_gst", 0)),
            _num(data_row.get("cgst_sgst_tally", data_row.get("tax_amount_tally", 0))) -
                _num(data_row.get("cgst_sgst_gst", data_row.get("tax_amount_gst", 0))),
            0,   # Cess diff
            (_num(data_row.get("taxable_value_tally", 0)) + _num(data_row.get("tax_amount_tally", 0))) -
                (_num(data_row.get("taxable_value_gst", 0)) + _num(data_row.get("tax_amount_gst", 0))),
            status,
            _safe(data_row.get("Reason", "")),
        ]

        # Row fill based on match status
        _status_fill = {
            "Matched":           _fill(CLR_ROW_MATCHED),
            "Partially Matched": _fill(CLR_ROW_PARTIAL),
            "In Books Only":     _fill(CLR_ROW_BOOKS),
            "In 2B Only":        _fill(CLR_ROW_2B),
        }
        row_fill = _status_fill.get(status)
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            if row_fill:
                cell.fill = row_fill


def _hdr(ws, row, col, value, bg_color, font_color=CLR_WHITE):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(bg_color)
    c.font = _font(bold=True, color=font_color)
    c.alignment = _center()
    c.border = _thin_border()
    return c


# ─────────────────────────────────────────────
# Supplier-Wise sheet builder
# ─────────────────────────────────────────────
def _build_supplier_wise_sheet(ws, supplier_rows, buyer_name, gstin, period,
                                total_suppliers, matched, partial, books_only, gst_only):
    """
    Mirrors 'Supplier Wise 2B':
      Row 1 – buyer info banner
      Row 2 – supplier-level count banner
      Row 3 – column group headers  (A..G plain, H..M Books, N..S GSTR 2B, T..X Diff)
      Row 4 – sub-column headers
      Row 5+ – data (one row per supplier)
    """
    _write_info_rows(ws, buyer_name, gstin, period, total_suppliers,
                     matched, partial, books_only, gst_only,
                     row1_label="Total Suppliers →")

    # ── Row 3: group headers ──────────────────
    static_cols = ["Supplier Name", "GSTIN", "Cancellation Date",
                   "Matched", "Partially Matched", "In 2B Only", "In Books Only"]
    for col_idx, label in enumerate(static_cols, start=1):
        ws.merge_cells(
            start_row=3, start_column=col_idx,
            end_row=4,   end_column=col_idx
        )
        _hdr(ws, 3, col_idx, label, CLR_SALMON)

    # Books H3:M3
    ws.merge_cells("H3:M3")
    _hdr(ws, 3, 8, "Books", CLR_STEEL)

    # GSTR 2B N3:S3
    ws.merge_cells("N3:S3")
    _hdr(ws, 3, 14, "GSTR 2B", CLR_STEEL)

    # Difference T3:X3
    ws.merge_cells("T3:X3")
    _hdr(ws, 3, 20, "Difference", CLR_STEEL)

    # ── Row 4: sub-column headers ─────────────
    sub = [
        # Books
        (8,  "No. of Records"),
        (9,  "Taxable Value"),
        (10, "IGST"),
        (11, "CGST + SGST"),
        (12, "Cess"),
        (13, "Invoice Value"),
        # GSTR 2B
        (14, "No. of Records"),
        (15, "Taxable Value"),
        (16, "IGST"),
        (17, "CGST + SGST"),
        (18, "Cess"),
        (19, "Invoice Value"),
        # Difference
        (20, "Taxable Value Diff."),
        (21, "IGST Diff."),
        (22, "CGST + SGST Diff."),
        (23, "Cess Diff."),
        (24, "Invoice Value Diff."),
    ]
    for col_idx, label in sub:
        _hdr(ws, 4, col_idx, label, CLR_SALMON)

    # ── Column widths ─────────────────────────
    widths = {1: 35, 2: 20, 4: 20, 5: 20, 6: 20, 7: 20,
              8: 20, 9: 20, 14: 20, 15: 20}
    for c_idx in range(1, 25):
        w = widths.get(c_idx, 15)
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    ws.freeze_panes = "B5"

    # ── Data rows ─────────────────────────────
    for srow in supplier_rows:
        r = ws.max_row + 1
        vals = [
            srow.get("supplier_name", ""),
            srow.get("gstin", ""),
            srow.get("cancellation_date", "-"),
            srow.get("matched", 0),
            srow.get("partially_matched", 0),
            srow.get("in_2b_only", 0),
            srow.get("in_books_only", 0),
            # Books
            srow.get("books_records", 0),
            round(srow.get("books_taxable", 0), 2),
            round(srow.get("books_igst", 0), 2),
            round(srow.get("books_cgst_sgst", 0), 2),
            0,  # Cess
            round(srow.get("books_invoice_value", 0), 2),
            # GSTR 2B
            srow.get("gst_records", 0),
            round(srow.get("gst_taxable", 0), 2),
            round(srow.get("gst_igst", 0), 2),
            round(srow.get("gst_cgst_sgst", 0), 2),
            0,  # Cess
            round(srow.get("gst_invoice_value", 0), 2),
            # Difference
            round(srow.get("books_taxable", 0) - srow.get("gst_taxable", 0), 2),
            round(srow.get("books_igst", 0) - srow.get("gst_igst", 0), 2),
            round(srow.get("books_cgst_sgst", 0) - srow.get("gst_cgst_sgst", 0), 2),
            0,  # Cess diff
            round(srow.get("books_invoice_value", 0) - srow.get("gst_invoice_value", 0), 2),
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            if r % 2 == 0:
                cell.fill = _fill(CLR_ROW_ALT)


# ─────────────────────────────────────────────
# Original parsing functions (unchanged)
# ─────────────────────────────────────────────

def parse_tally_file(file_path):
    raw = pd.read_excel(file_path, nrows=20, header=None)
    header_row = None

    for i in range(raw.shape[0]):
        row_str = ' '.join([str(x).lower() for x in raw.iloc[i] if pd.notna(x)])
        if any(keyword in row_str for keyword in ['voucher no', 'bill no', 'invoice number', 'invoice no']):
            header_row = i
            break

    if header_row is None:
        raise ValueError("Could not find Tally header row (looking for 'Voucher No', 'Bill No', or 'Invoice Number')")

    file_path.seek(0)
    df = pd.read_excel(file_path, header=header_row)
    raw_columns = [str(c).strip() for c in df.columns]

    col_map = {}
    for col in df.columns:
        col_lower = str(col).lower().strip()
        if 'invoice number' in col_lower or 'invoice no' in col_lower or 'voucher no' in col_lower or 'bill no' in col_lower:
            if 'invoice_no' not in col_map.values():
                col_map[col] = 'invoice_no'
        elif col_lower == 'date' or 'invoice date' in col_lower:
            col_map[col] = 'invoice_date'
        elif 'gstin' in col_lower:
            col_map[col] = 'gstin'
        elif 'party' in col_lower:
            col_map[col] = 'party_name_tally'
        elif 'gross minus discount' in col_lower or 'taxable value' in col_lower:
            col_map[col] = 'taxable_value'
        elif 'gst amount' in col_lower or 'tax amount' in col_lower or 'total tax' in col_lower:
            col_map[col] = 'tax_amount'
        elif 'integrated tax' in col_lower or col_lower == 'igst':
            col_map[col] = 'igst'
        elif 'central tax' in col_lower or col_lower == 'cgst':
            col_map[col] = 'cgst'
        elif 'state/ut tax' in col_lower or 'state tax' in col_lower or col_lower == 'sgst':
            col_map[col] = 'sgst'

    df.rename(columns=col_map, inplace=True)

    if 'tax_amount' not in df.columns:
        tax_sum = pd.Series(0.0, index=df.index)
        found_taxes = False
        for t_col in ['igst', 'cgst', 'sgst']:
            if t_col in df.columns:
                found_taxes = True
                temp = df[t_col].astype(str).str.replace(',', '', regex=False).str.replace('₹', '', regex=False).str.strip()
                tax_sum += pd.to_numeric(temp, errors='coerce').fillna(0)
        if found_taxes:
            df['tax_amount'] = tax_sum

    if 'invoice_no' not in df.columns:
        df['invoice_no'] = ''

    df['invoice_no'] = df['invoice_no'].astype(str).str.strip().replace(['nan', 'NA'], '')

    is_primary = pd.Series(False, index=df.index)
    if 'invoice_date' in df.columns:
        is_primary |= df['invoice_date'].astype(str).str.strip().replace(['nan', 'NA', 'NaT', 'None'], '') != ''
    if 'party_name_tally' in df.columns:
        is_primary |= df['party_name_tally'].astype(str).str.strip().replace(['nan', 'NA', 'None'], '') != ''

    blank_primary_mask = is_primary & (df['invoice_no'] == '')
    df.loc[blank_primary_mask, 'invoice_no'] = ['NO_INV_' + str(i) for i in df.index[blank_primary_mask]]
    df['invoice_no'] = df['invoice_no'].replace('', np.nan).ffill()

    if 'gstin' in df.columns:
        df['gstin'] = df['gstin'].astype(str).str.strip().str.upper()
        df['gstin'] = df['gstin'].replace(['NAN', 'NA', 'NONE', ''], np.nan).ffill().fillna('UNREGISTERED')

    if 'party_name_tally' in df.columns:
        df['party_name_tally'] = df['party_name_tally'].astype(str).str.strip().replace(['nan', 'NA', 'None', ''], np.nan).ffill()

    if 'invoice_date' in df.columns:
        df['invoice_date'] = df['invoice_date'].astype(str).str.strip().replace(['nan', 'NA', 'NaT', 'None', ''], np.nan).ffill()

    product_cols = [c for c in df.columns if 'product' in str(c).lower() or 'other account' in str(c).lower()]
    round_off_mask = pd.Series(False, index=df.index)
    for col in product_cols:
        round_off_mask |= df[col].astype(str).str.lower().str.contains('round off', na=False)

    special_rows = df[round_off_mask].copy()
    agg_cols = []
    for col in ['taxable_value', 'tax_amount']:
        if col in special_rows.columns:
            special_rows[col] = special_rows[col].astype(str).str.replace(',', '', regex=False).str.replace('₹', '', regex=False).str.strip()
            special_rows[col] = pd.to_numeric(special_rows[col], errors='coerce').fillna(0)
            agg_cols.append(col)

    if agg_cols and not special_rows.empty:
        special_agg = special_rows.groupby('invoice_no')[agg_cols].sum()
    else:
        special_agg = pd.DataFrame()

    df = df[~round_off_mask]

    if not special_agg.empty:
        adjustments = special_agg.reset_index()
        if 'gstin' in df.columns:
            adjustments['gstin'] = adjustments['invoice_no'].map(df.groupby('invoice_no')['gstin'].first())
        if 'party_name_tally' in df.columns:
            adjustments['party_name_tally'] = adjustments['invoice_no'].map(df.groupby('invoice_no')['party_name_tally'].first())
        if 'invoice_date' in df.columns:
            adjustments['invoice_date'] = adjustments['invoice_no'].map(df.groupby('invoice_no')['invoice_date'].first())
        df = pd.concat([df, adjustments], ignore_index=True)

    df = df[df['invoice_no'].notna()]
    df = df[~df['invoice_no'].astype(str).str.contains('Total', na=False, case=False)]
    return df, raw_columns


def aggregate_tally(df):
    if 'invoice_no' not in df.columns:
        df['invoice_no'] = ''
    if 'gstin' not in df.columns:
        df['gstin'] = ''

    df['invoice_no'] = df['invoice_no'].astype(str).str.strip()
    df['gstin'] = df['gstin'].astype(str).str.strip().str.upper()

    for col in ['taxable_value', 'tax_amount']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(',', '', regex=False).str.replace('₹', '', regex=False).str.strip()
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        else:
            df[col] = 0.0

    if 'invoice_date' in df.columns:
        df['invoice_date'] = pd.to_datetime(df['invoice_date'], errors='coerce', dayfirst=True)

    agg_funcs = {'taxable_value': 'sum', 'tax_amount': 'sum'}
    if 'invoice_date' in df.columns:
        agg_funcs['invoice_date'] = 'first'
    if 'party_name_tally' in df.columns:
        agg_funcs['party_name_tally'] = 'first'

    agg = df.groupby(['invoice_no', 'gstin']).agg(agg_funcs).reset_index()

    if 'invoice_date' in agg.columns:
        agg['invoice_date'] = agg['invoice_date'].dt.strftime('%d/%m/%Y')

    # Bug-fix: add normalised key for separator-agnostic matching
    agg['invoice_no_norm'] = agg['invoice_no'].apply(_normalize_inv)

    return agg


def parse_gstr2b_file(file_path):
    file_path.seek(0)
    xls = pd.ExcelFile(file_path)
    sheet_name = 'B2B' if 'B2B' in xls.sheet_names else xls.sheet_names[0]

    raw = pd.read_excel(file_path, sheet_name=sheet_name, nrows=20, header=None)
    header_row = None

    for i in range(raw.shape[0]):
        row_str = ' '.join([str(x).lower() for x in raw.iloc[i] if pd.notna(x)])
        if ('invoice number' in row_str or 'invoice no' in row_str) and 'gstin' in row_str:
            header_row = i
            break

    if header_row is None:
        header_row = 2

    file_path.seek(0)
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)

    col_map = {}
    for col in df.columns:
        col_lower = str(col).lower().strip().replace('\n', ' ')
        if 'invoice number' in col_lower or 'invoice no' in col_lower:
            col_map[col] = 'invoice_no'
        elif 'gstin' in col_lower:
            if 'gstin' not in col_map.values():
                col_map[col] = 'gstin'
        elif 'trade/legal name' in col_lower or 'trade name' in col_lower or 'legal name' in col_lower:
            col_map[col] = 'party_name_gst'
        elif 'invoice date' in col_lower:
            col_map[col] = 'invoice_date'
        elif 'taxable value' in col_lower:
            col_map[col] = 'taxable_value'
        elif 'integrated tax' in col_lower or col_lower == 'igst':
            col_map[col] = 'igst'
        elif 'central tax' in col_lower or col_lower == 'cgst':
            col_map[col] = 'cgst'
        elif 'state/ut tax' in col_lower or 'state tax' in col_lower or col_lower == 'sgst':
            col_map[col] = 'sgst'
        elif 'total tax' in col_lower or 'tax amount' in col_lower:
            col_map[col] = 'tax_amount_native'

    df.rename(columns=col_map, inplace=True)

    if 'invoice_no' not in df.columns:
        df['invoice_no'] = ''
    if 'gstin' not in df.columns:
        df['gstin'] = ''

    df['gstin'] = df['gstin'].astype(str).str.strip().str.upper()
    df['invoice_no'] = df['invoice_no'].astype(str).str.strip()

    for col in ['igst', 'cgst', 'sgst', 'taxable_value', 'tax_amount_native']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(',', '', regex=False).str.replace('₹', '', regex=False).str.replace('- 0', '0', regex=False).str.replace('-0', '0', regex=False).str.strip()
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        else:
            df[col] = 0.0

    if 'tax_amount_native' in df.columns and df['tax_amount_native'].sum() > 0:
        df['tax_amount'] = df['tax_amount_native']
    else:
        df['tax_amount'] = df['igst'] + df['cgst'] + df['sgst']

    if 'invoice_date' in df.columns:
        df['invoice_date'] = pd.to_datetime(df['invoice_date'], dayfirst=True, errors='coerce').dt.strftime('%d/%m/%Y')

    needed = ['gstin', 'invoice_no', 'invoice_date', 'taxable_value', 'tax_amount', 'party_name_gst']
    df = df[[c for c in needed if c in df.columns]]
    df = df.dropna(subset=['invoice_no'])
    df = df[df['invoice_no'] != 'nan']
    df = df[df['invoice_no'] != '']

    # Bug-fix: add normalised key for separator-agnostic matching
    df['invoice_no_norm'] = df['invoice_no'].apply(_normalize_inv)

    return df


# ─────────────────────────────────────────────
# Main reconcile function – now produces the
# sample-format XLSX output
# ─────────────────────────────────────────────

def reconcile(file_tally, file_gstr2b,
              buyer_name="", gstin="", period=""):
    warnings_list = []

    # ── 1. Parse & aggregate ──────────────────
    df_tally_raw, tally_raw_columns = parse_tally_file(file_tally)
    tally_df = aggregate_tally(df_tally_raw)
    gst_df = parse_gstr2b_file(file_gstr2b)

    if 'Bill No' not in tally_raw_columns and 'bill no' not in [c.lower() for c in tally_raw_columns]:
        warnings_list.append("⚠️ Tally file missing 'Bill No' column. Re-export from Tally with Bill No/Supplier Invoice No column for accurate matching.")

    # ── 2. Date helpers ───────────────────────
    if 'invoice_date' in tally_df.columns:
        tally_df['invoice_date_dt'] = pd.to_datetime(tally_df['invoice_date'], errors='coerce', dayfirst=True)
    else:
        tally_df['invoice_date_dt'] = pd.NaT

    if 'invoice_date' in gst_df.columns:
        gst_df['invoice_date_dt'] = pd.to_datetime(gst_df['invoice_date'], errors='coerce', dayfirst=True)
    else:
        gst_df['invoice_date_dt'] = pd.NaT

    # ── 3. Match on normalised invoice_no (/ treated same as -) ─────────
    #
    # THE FIX: merge on `invoice_no_norm` (canonical key) instead of the
    # raw `invoice_no` string.  Both DataFrames carry their ORIGINAL invoice
    # number in `invoice_no`; after the merge those become `invoice_no_tally`
    # and `invoice_no_gst`.  We then restore a plain `invoice_no` column
    # (preferring the Tally original) so all downstream code stays unchanged.
    merged = pd.merge(tally_df, gst_df, on='invoice_no_norm', how='outer',
                      indicator=True, suffixes=('_tally', '_gst'))

    # Restore backward-compat `invoice_no` column for all downstream consumers
    if 'invoice_no_tally' in merged.columns and 'invoice_no_gst' in merged.columns:
        merged['invoice_no'] = (
            merged['invoice_no_tally']
            .combine_first(merged['invoice_no_gst'])
            .combine_first(merged['invoice_no_norm'])
        )
    elif 'invoice_no_tally' in merged.columns:
        merged['invoice_no'] = merged['invoice_no_tally'].combine_first(merged['invoice_no_norm'])
    elif 'invoice_no_gst' in merged.columns:
        merged['invoice_no'] = merged['invoice_no_gst'].combine_first(merged['invoice_no_norm'])

    matched          = merged[merged['_merge'] == 'both'].copy()
    missing_in_gst   = merged[merged['_merge'] == 'left_only'].copy()
    missing_in_tally = merged[merged['_merge'] == 'right_only'].copy()

    # ── 4. Fallback match (gstin + date ±3d + value ±1%) ────
    missing_in_gst_drop   = []
    missing_in_tally_drop = []
    fallback_matches      = []

    if not missing_in_gst.empty and not missing_in_tally.empty:
        for gst_idx, gst_row in missing_in_tally.iterrows():
            g_gstin = gst_row['gstin_gst']
            g_date  = gst_row['invoice_date_dt_gst']
            g_val   = gst_row['taxable_value_gst']

            candidates = missing_in_gst[
                (missing_in_gst['gstin_tally'] == g_gstin) &
                (~missing_in_gst.index.isin(missing_in_gst_drop))
            ]
            if candidates.empty:
                continue

            if pd.notnull(g_date):
                candidates = candidates[
                    (candidates['invoice_date_dt_tally'].notnull()) &
                    (abs((candidates['invoice_date_dt_tally'] - g_date).dt.days) <= 3)
                ]
            if candidates.empty:
                continue

            candidates = candidates[
                abs(candidates['taxable_value_tally'] - g_val) <= (0.01 * abs(g_val) + 0.01)
            ]
            if not candidates.empty:
                t_idx = candidates.index[0]
                t_row = candidates.loc[t_idx]
                fallback_matches.append({
                    'invoice_no':           t_row.get('invoice_no_tally', t_row.get('invoice_no')),
                    'gstin_tally':          t_row['gstin_tally'],
                    'invoice_date_tally':   t_row.get('invoice_date_tally'),
                    'taxable_value_tally':  t_row['taxable_value_tally'],
                    'tax_amount_tally':     t_row['tax_amount_tally'],
                    'party_name_tally':     t_row.get('party_name_tally'),
                    'gstin_gst':            g_gstin,
                    'invoice_date_gst':     gst_row.get('invoice_date_gst'),
                    'taxable_value_gst':    g_val,
                    'tax_amount_gst':       gst_row['tax_amount_gst'],
                    'invoice_no_gst':       gst_row.get('invoice_no_gst', gst_row.get('invoice_no')),
                    'party_name_gst':       gst_row.get('party_name_gst'),
                    '_merge':               'fallback_match',
                })
                missing_in_gst_drop.append(t_idx)
                missing_in_tally_drop.append(gst_idx)

    missing_in_gst   = missing_in_gst.drop(index=missing_in_gst_drop)
    missing_in_tally = missing_in_tally.drop(index=missing_in_tally_drop)

    if fallback_matches:
        matched = pd.concat([matched, pd.DataFrame(fallback_matches)], ignore_index=True)

    # ── 5. Classify matched as Matched / Mismatched ──
    final_matched    = []
    final_mismatched = []

    for _, row in matched.iterrows():
        t_val = float(row.get('taxable_value_tally') or 0)
        g_val = float(row.get('taxable_value_gst')   or 0)
        t_tax = float(row.get('tax_amount_tally')    or 0)
        g_tax = float(row.get('tax_amount_gst')      or 0)

        val_diff = abs(t_val - g_val)
        tax_diff = abs(t_tax - g_tax)

        mismatches = []
        if val_diff > 1.0:
            mismatches.append(f"Taxable Value (GST: {g_val}, Tally: {t_val})")
        if tax_diff > 1.0:
            mismatches.append(f"Tax Amount (GST: {g_tax}, Tally: {t_tax})")

        row_dict = row.to_dict()
        if not mismatches:
            row_dict['Match Status'] = 'Matched'
            row_dict['Reason'] = ('Exact Match'    if row.get('_merge') == 'both'
                                  else 'Fallback Match') if val_diff == 0 and tax_diff == 0 \
                                  else 'Matched (Rounding Difference)'
            final_matched.append(row_dict)
        else:
            row_dict['Match Status'] = 'Partially Matched'
            row_dict['Reason'] = " | ".join(mismatches)
            final_mismatched.append(row_dict)

    df_matched    = pd.DataFrame(final_matched)
    df_mismatched = pd.DataFrame(final_mismatched)

    missing_in_gst['Match Status']   = 'In Books Only'
    missing_in_gst['Reason']         = ''
    missing_in_tally['Match Status'] = 'In 2B Only'
    missing_in_tally['Reason']       = ''

    # Drop helper cols
    drop_cols = ['invoice_date_dt_tally', 'invoice_date_dt_gst', 'invoice_date_dt',
                 '_merge', 'invoice_no_norm', 'invoice_no_tally', 'invoice_no_gst']
    for df in [df_matched, df_mismatched, missing_in_gst, missing_in_tally]:
        if not df.empty:
            df.drop(columns=[c for c in drop_cols if c in df.columns], inplace=True, errors='ignore')

    # ── 6. Build invoice-level records list ───
    def _to_records(df):
        return df.to_dict('records') if not df.empty else []

    all_records = (
        _to_records(df_matched) +
        _to_records(df_mismatched) +
        _to_records(missing_in_gst) +
        _to_records(missing_in_tally)
    )

    # ── 7. Build Supplier-Wise summary ────────
    supplier_map = {}

    def _add_supplier(gstin_key, name, rec, side):
        if gstin_key not in supplier_map:
            supplier_map[gstin_key] = {
                "supplier_name": name or gstin_key,
                "gstin": gstin_key,
                "cancellation_date": "-",
                "matched": 0, "partially_matched": 0,
                "in_2b_only": 0, "in_books_only": 0,
                "books_records": 0, "books_taxable": 0,
                "books_igst": 0, "books_cgst_sgst": 0,
                "books_invoice_value": 0,
                "gst_records": 0, "gst_taxable": 0,
                "gst_igst": 0, "gst_cgst_sgst": 0,
                "gst_invoice_value": 0,
            }
        s = supplier_map[gstin_key]

        tv_t = float(rec.get('taxable_value_tally') or 0)
        ta_t = float(rec.get('tax_amount_tally')    or 0)
        tv_g = float(rec.get('taxable_value_gst')   or 0)
        ta_g = float(rec.get('tax_amount_gst')      or 0)
        ms   = rec.get('Match Status', '')

        if ms == 'Matched':
            s['matched'] += 1
        elif ms == 'Partially Matched':
            s['partially_matched'] += 1
        elif ms == 'In Books Only':
            s['in_books_only'] += 1
        elif ms == 'In 2B Only':
            s['in_2b_only'] += 1

        if side in ('both', 'books'):
            s['books_records']       += 1
            s['books_taxable']       += tv_t
            s['books_cgst_sgst']     += ta_t
            s['books_invoice_value'] += tv_t + ta_t

        if side in ('both', 'gst'):
            s['gst_records']       += 1
            s['gst_taxable']       += tv_g
            s['gst_cgst_sgst']     += ta_g
            s['gst_invoice_value'] += tv_g + ta_g

    for rec in _to_records(df_matched):
        g = rec.get('gstin_tally') or rec.get('gstin_gst', '')
        n = rec.get('party_name_tally') or rec.get('party_name_gst', '')
        _add_supplier(g, n, rec, 'both')

    for rec in _to_records(df_mismatched):
        g = rec.get('gstin_tally') or rec.get('gstin_gst', '')
        n = rec.get('party_name_tally') or rec.get('party_name_gst', '')
        _add_supplier(g, n, rec, 'both')

    for rec in _to_records(missing_in_gst):
        g = rec.get('gstin_tally', '')
        n = rec.get('party_name_tally', '')
        _add_supplier(g, n, rec, 'books')

    for rec in _to_records(missing_in_tally):
        g = rec.get('gstin_gst', '')
        n = rec.get('party_name_gst', '')
        _add_supplier(g, n, rec, 'gst')

    supplier_rows = sorted(supplier_map.values(), key=lambda x: x['supplier_name'])

    # ── 8. Count totals ───────────────────────
    total_matched   = len(df_matched)
    total_partial   = len(df_mismatched)
    total_bks_only  = len(missing_in_gst)
    total_gst_only  = len(missing_in_tally)
    total_docs      = total_matched + total_partial + total_bks_only + total_gst_only
    total_suppliers = len(supplier_map)

    summary_metrics = {
        'Total Records':    total_docs,
        'Matched':          total_matched,
        'Mismatched':       total_partial,
        'Missing in GST':   total_bks_only,
        'Missing in Tally': total_gst_only,
    }

    # ── 9. Write the output workbook ──────────
    output = io.BytesIO()
    wb = Workbook()

    # Remove the default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Sheet 1: Supplier Wise 2B
    ws_sup = wb.create_sheet("Supplier Wise 2B")
    _build_supplier_wise_sheet(
        ws_sup, supplier_rows,
        buyer_name, gstin, period,
        total_suppliers, total_matched, total_partial, total_bks_only, total_gst_only
    )

    # Sheet 2: Invoice Wise 2B (B2B_CNDN_ECO)
    ws_inv = wb.create_sheet("Invoice Wise 2B(B2B_CNDN_ECO)")
    _build_invoice_wise_sheet(
        ws_inv, all_records,
        buyer_name, gstin, period,
        total_matched, total_partial, total_bks_only, total_gst_only
    )

    # Sheet 3: IMPG (empty placeholder matching the sample)
    ws_impg = wb.create_sheet("Invoice Wise 2B(IMPG_IMPGSEZ)")
    _write_info_rows(ws_impg, buyer_name, gstin, period, 0, 0, 0, 0, 0)

    # Sheet 4: ISD (empty placeholder)
    ws_isd = wb.create_sheet("Invoice Wise 2B(ISD)")
    _write_info_rows(ws_isd, buyer_name, gstin, period, 0, 0, 0, 0, 0,
                     row1_label="Total Documents →")

    wb.save(output)
    output.seek(0)

    return summary_metrics, output, warnings_list, missing_in_gst


def generate_supplier_defaulters_list(missing_in_gst_df):
    """Generate a professional Supplier Defaulters List for emails/follow-ups"""
    if missing_in_gst_df.empty:
        return None

    defaulter_cols = {
        'invoice_no':          'Voucher No (Tally)',
        'gstin_tally':         'Supplier GSTIN',
        'party_name_tally':    'Supplier Name',
        'invoice_date_tally':  'Invoice Date',
        'taxable_value_tally': 'Taxable Value (₹)',
        'tax_amount_tally':    'Tax Amount (₹)',
    }

    defaulters = missing_in_gst_df[[col for col in defaulter_cols if col in missing_in_gst_df.columns]].copy()
    defaulters.rename(columns=defaulter_cols, inplace=True)

    if 'Taxable Value (₹)' in defaulters.columns and 'Tax Amount (₹)' in defaulters.columns:
        defaulters['Total Invoice Value (₹)'] = (
            defaulters['Taxable Value (₹)'].fillna(0) +
            defaulters['Tax Amount (₹)'].fillna(0)
        )

    defaulters['Action Required'] = 'Supplier must file GSTR-1'

    if 'Supplier Name' in defaulters.columns:
        defaulters.sort_values(by=['Supplier Name', 'Invoice Date'], inplace=True)

    defaulters.reset_index(drop=True, inplace=True)

    output = io.BytesIO()
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Supplier Defaulters"

    salmon_fill  = _fill(CLR_SALMON)
    steel_fill   = _fill(CLR_STEEL)
    white_font   = _font(bold=True, color=CLR_WHITE)
    title_font   = Font(bold=True, size=14, color="1F4E78")
    bold_font    = _font(bold=True)

    # Title
    ws2.cell(row=1, column=1, value="SUPPLIER DEFAULTERS LIST - GSTR-1 NOT FILED").font = title_font

    # Summary row (row 2)
    summary_vals = {
        'Voucher No (Tally)':      'SUMMARY',
        'Supplier GSTIN':          '',
        'Supplier Name':           f'Total Missing Invoices: {len(defaulters)}',
        'Invoice Date':            '',
        'Taxable Value (₹)':       defaulters['Taxable Value (₹)'].sum() if 'Taxable Value (₹)' in defaulters.columns else 0,
        'Tax Amount (₹)':          defaulters['Tax Amount (₹)'].sum()    if 'Tax Amount (₹)' in defaulters.columns else 0,
        'Total Invoice Value (₹)': defaulters['Total Invoice Value (₹)'].sum() if 'Total Invoice Value (₹)' in defaulters.columns else 0,
        'Action Required':         '',
    }
    for col_idx, val in enumerate(summary_vals.values(), start=1):
        c = ws2.cell(row=2, column=col_idx, value=val)
        c.fill  = _fill("FFFFC000")
        c.font  = bold_font
        c.border = _thin_border()

    # Column headers (row 3)
    for col_idx, col_name in enumerate(defaulters.columns, start=1):
        c = ws2.cell(row=3, column=col_idx, value=col_name)
        c.fill      = salmon_fill
        c.font      = white_font
        c.alignment = _center()
        c.border    = _thin_border()

    # Data rows
    for row_idx, (_, data_row) in enumerate(defaulters.iterrows(), start=4):
        row_fill = _fill("FFE8F5E9") if row_idx % 2 == 0 else _fill(CLR_WHITE)
        for col_idx, val in enumerate(data_row.values, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = _thin_border()

    # Column widths
    col_widths = [15, 20, 35, 12, 18, 15, 20, 30]
    for i, w in enumerate(col_widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb2.save(output)
    output.seek(0)
    return output